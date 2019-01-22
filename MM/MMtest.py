from openpyxl import load_workbook
from openpyxl import Workbook
import numpy
import operator

wb = load_workbook(filename = '/Users/JKTechnical/Desktop/MM/Rollercoaster.xlsx')
ws = wb['RollerCoasterData']

def distance(lamList):
	
	#Weight
	wH = .076664378 #Height 
	wS = .018044793 #Speed  
	wL = .065437140 #Length 
	wI = .347962998 #Inversion 
	wD = .021563261 #Drop 
	wT = .251311299 #Duration 
	wA = .219016131 #Angle 

	wList = [wH, wS, wL, wI, wD, wT, wA]

	dist = 0
	for lam in range(len(lamList)):
		dist = dist + (abs(lamList[lam] - 1) * wList[lam])

	return dist

# - - - - - - - - - - - - - - - - - - - - - - - - - - - 

def algorithm(name, *args):
	
	name = name
	
	#overall averages
	mH = 135.2994147
	mS = 59.6774248
	mL = 3149.9392272
	mI = 2.223333
	mD = 153.1843972
	mT = 127
	mA = 74.5711111

	#Ideal values
	iH = 216.361797  #Height
	iS = 78.994382 	 #Speed
	iL = 4856.625843 #Length
	iI = 0.921348    #Inversions
	iD = 192.884727  #Drop
	iT = 102.067415  #Duration
	iA = 62.455056   #Angle

	#iterates over each attribute of the coaster
	print('Number of arguments is {}'.format(len(args)))
	for ind in range(len(args)):
 		
 		if type(args[ind]) != float and type(args[ind])!= int:
 			
 			return (name, 5)

	 	else:								
	 		#otherwise set the actual to the iput
	 		if ind == 0: aH = args[ind]; print('H assigned val {}'.format(args[ind]))
	 		if ind == 1: aS = args[ind]; print('S assigned val {}'.format(args[ind]))
	 		if ind == 2: aL = args[ind]; print('L assigned val {}'.format(args[ind]))
	 		if ind == 3: aI = args[ind]; print('I assigned val {}'.format(args[ind]))
	 		if ind == 4: aD = args[ind]; print('D assigned val {}'.format(args[ind]))
	 		if ind == 5: aT = args[ind]; print('T assigned val {}'.format(args[ind]))
	 		if ind == 6: aA = args[ind]; print('A assigned val {}'.format(args[ind]))

	# Assigning the lambda values for each attribute
	gH = aH / iH
	gS = aS / iH 
	gL = aL / iL
	gI = aI / iI
	gD = aD / iD
	gT = aT / iT
	gA = aA / iA

	#puts all lambdas in a single list to make things easier
	lamList = [gH, gS, gL, gI, gD, gT, gA]
		
	dist = distance(lamList)
	print('name is {}'.format(name))
	print('dist {}'.format(dist))
	
	return (name, dist)

# - - - - - - - - - - - - - - - - - - - - - - - - - - - 


if __name__ == '__main__':
	
	rollerRank = {}
	rownum = 1
	
	for row in ws.iter_rows(min_row=2, max_col=19, max_row=301):
		colnum = 0
		rownum = rownum + 1
		
		#Initializing actual values
		aN = 'SampleName'
		aH = 0 
		aS = 0
		aL = 0
		aI = 0
		aD = 0
		aT = 0
		aA = 0

		#get values
		for cell in row:
			colnum = colnum + 1
			
			#Title 
			if colnum == 1: 
				aN = cell.value
			
			'''
			#Park 
			if colnum == 2:
				pass
			#Material
			if colnum == 7:
				pass
			#Type
			if colnum == 8:
				pass
			'''
			
			#Height
			if colnum == 11:
				if type(cell.value) == float or type(cell.value) == int:
					aH = cell.value
				else:
					print(cell.value)
					aH = 'h'
			
			#Speed
			if colnum == 12:
				if type(cell.value) == float or type(cell.value) == int:
					aS = cell.value
				else:
					print(cell.value)
					aS = 's'

			#Length
			if colnum == 13:
				if type(cell.value) == float or type(cell.value) == int:
					aL = cell.value
				else:
					print(cell.value)
					aL = 'L'
		
			#Inversions
			if colnum == 15:
				if type(cell.value) == float or type(cell.value) == int:
					aI = cell.value
				else:
					print(cell.value)
					aI = 'I'

			#Drop
			if colnum == 16:
				if type(cell.value) == float or type(cell.value) == int:
					aD = cell.value
				else:
					print(cell.value)
					aD = 'D'

			#Duration (secs)
			if colnum == 17:
				if type(cell.value) == float or type(cell.value) == int:
					aT = cell.value
				else:
					print(cell.value)
					aT = 'T'

			#Angle
			if colnum == 19:
				if type(cell.value) == float or type(cell.value) == int:
					aA = cell.value
					print('fuck god')
				else:
					print('Faulty angle {}'.format(cell.value))
					aA = 'A'

			print('colnum is {}'.format(colnum))
			print('row is {}'.format(rownum))
			print('\n\n')


		name, dist 			= algorithm(aN, aH, aS, aL, aI, aD, aT, aA)
		rollerRank[name]    = dist
		orderDict			= sorted(rollerRank.items(), key=operator.itemgetter(1))
		for i in orderDict:
			print(i)
			print('\n')



